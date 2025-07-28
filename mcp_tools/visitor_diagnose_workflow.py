"""매장 방문객수 진단 보고서 작성을 도와주는 워크플로우"""

import json
import os
import sys
from typing import Dict, Any, List, Union
from fastmcp import FastMCP  # FastMCP 툴 서버용

from dotenv import load_dotenv
from langchain.schema import BaseOutputParser
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, END, START

from base_workflow import BaseWorkflow, BaseState


class VisitorDiagnoseState(BaseState):
    """방문객 진단 워크플로우 전용 상태 - BaseState 확장"""
    store_name: str  # 단일 매장명
    period: str
    visitor_data: Dict[str, Any]  # 단일 매장 데이터
    raw_answer: str  # diagnose_avg_in의 원시 응답
    metric_dict: Dict[str, Any]  # 파싱된 메트릭 데이터
    placements: List[Dict[str, Any]]  # 엑셀 셀 배치 정보
    final_result: str  # 최종 결과
    design_spec: List[Dict[str, Any]]  # 디자인 스타일 placements


class VisitorDiagnoseWorkflow(BaseWorkflow[VisitorDiagnoseState]):
    """방문객 진단 워크플로우 클래스 - BaseWorkflow 상속"""

    def __init__(self):
        super().__init__(workflow_name="visitor_diagnose")
        
        # 환경변수 로드 및 LLM 설정
        load_dotenv()
        self.llm = ChatOpenAI(model="gpt-4o", temperature=0)
        
        # 워크플로우 그래프
        self.workflow_app = self._build_workflow()

    def run(
        self, 
        user_prompt: str, 
        store_name: str, 
        start_date: str,
        end_date: str
    ) -> str:
        """Agent가 호출하는 방문객 진단 워크플로우 실행"""
        self.logger.info(f"워크플로우 실행: {store_name} ({start_date}~{end_date})")
        
        # 초기 상태 설정
        initial_state = VisitorDiagnoseState(
            store_name=store_name,
            period=f"{start_date}~{end_date}",
            visitor_data={},
            raw_answer="",
            metric_dict={},
            placements=[],
            final_result="",
            design_spec=[]
        )
        
        # 워크플로우 실행
        result = self.workflow_app.invoke(initial_state)
        
        return result.get("final_result", "워크플로우 실행 완료")

    def _build_workflow(self) -> StateGraph:
        """LangGraph 워크플로우를 구성합니다."""
        builder = StateGraph(VisitorDiagnoseState)
        
        # 노드 추가
        builder.add_node("fetch", self._query_db_node)
        builder.add_node("parse", self._parse_node)
        builder.add_node("map", self._map_to_excel_node)
        builder.add_node("update", self._update_excel_node)
        
        # 엣지 추가 (순차 실행)
        builder.add_edge(START, "fetch")
        builder.add_edge("fetch", "parse")
        builder.add_edge("parse", "map")
        builder.add_edge("map", "update")
        builder.add_edge("update", END)
        
        return builder.compile()

    # ----------------- 노드 구현 -----------------
    def _query_db_node(self, state: VisitorDiagnoseState) -> VisitorDiagnoseState:
        """diagnose_avg_in 로직을 직접 구현하여 DB에서 데이터 조회"""
        start, end = state["period"].split("~")
        
        # ClickHouse 클라이언트 생성
        client = self._create_clickhouse_client()
        if not client:
            state["raw_answer"] = "데이터베이스 연결 실패"
            return state
            
        # mcp_diagnose.py의 diagnose_avg_in 쿼리를 그대로 사용
        query = f"""
        WITH
df AS (
    SELECT
        li.date AS visit_date,
        li.timestamp,
        li.person_seq AS visitor_id,
        if(toDayOfWeek(li.date) IN (1,2,3,4,5), 'weekday', 'weekend')                     AS day_type,
        multiIf(
            toHour(li.timestamp) IN (22,23,0,1), '22-01',
            toHour(li.timestamp) BETWEEN 2  AND 5 , '02-05',
            toHour(li.timestamp) BETWEEN 6  AND 9 , '06-09',
            toHour(li.timestamp) BETWEEN 10 AND 13, '10-13',
            toHour(li.timestamp) BETWEEN 14 AND 17, '14-17',
            '18-21'
        ) AS time_range,
        multiIf(
            dt.age BETWEEN 0  AND  9 , '10대 미만',
            dt.age BETWEEN 10 AND 19, '10대',
            dt.age BETWEEN 20 AND 29, '20대',
            dt.age BETWEEN 30 AND 39, '30대',
            dt.age BETWEEN 40 AND 49, '40대',
            dt.age BETWEEN 50 AND 59, '50대',
            dt.age >= 60           , '60대 이상',
            'Unknown'
        ) AS age_group,
        if(dt.gender = '0', '남성', if(dt.gender='1','여성','Unknown'))                   AS gender
    FROM line_in_out_individual li
    LEFT JOIN detected_time dt ON li.person_seq = dt.person_seq
    LEFT JOIN line          l  ON li.triggered_line_id = l.id
    WHERE li.date BETWEEN '{start}' AND '{end}'
      AND li.is_staff = 0
      AND li.in_out   = 'IN'
      AND l.entrance  = 1
),
daily_all     AS (SELECT visit_date, uniqExact(visitor_id) AS ucnt FROM df GROUP BY visit_date),
avg_all       AS (SELECT toUInt64(round(avg(ucnt))) AS avg_cnt FROM daily_all),
daily_dayType AS (SELECT visit_date, day_type, uniqExact(visitor_id) AS ucnt FROM df GROUP BY visit_date, day_type),
avg_dayType   AS (SELECT day_type, toUInt64(round(avg(ucnt))) AS avg_cnt FROM daily_dayType GROUP BY day_type),
daily_gender  AS (SELECT visit_date, gender, uniqExact(visitor_id) AS ucnt FROM df GROUP BY visit_date, gender),
avg_gender    AS (SELECT gender, toUInt64(round(avg(ucnt))) AS avg_cnt FROM daily_gender GROUP BY gender),
daily_age     AS (SELECT visit_date, age_group, uniqExact(visitor_id) AS ucnt FROM df GROUP BY visit_date, age_group),
avg_age       AS (SELECT age_group, toUInt64(round(avg(ucnt))) AS avg_cnt FROM daily_age GROUP BY age_group),
rank_age      AS (SELECT *, row_number() OVER (ORDER BY avg_cnt DESC) AS rk FROM avg_age),
overall_cnt   AS (SELECT avg_cnt AS cnt FROM avg_all),
range_cnt AS (
    SELECT day_type, time_range, uniqExact(visitor_id) AS visit_cnt
    FROM df GROUP BY day_type, time_range
),
tot_cnt   AS (SELECT day_type, sum(visit_cnt) AS total_cnt FROM range_cnt GROUP BY day_type),
range_pct AS (
    SELECT r.day_type, r.time_range, r.visit_cnt,
           toUInt64(round(r.visit_cnt / t.total_cnt * 100)) AS pct
    FROM range_cnt r JOIN tot_cnt t USING(day_type)
),
rank_slot AS (
    SELECT *, row_number() OVER (PARTITION BY day_type ORDER BY pct DESC) AS rk
    FROM range_pct
),
final AS (
    SELECT '일평균' AS section, '전체' AS label,
           avg_cnt AS value_cnt, CAST(NULL AS Nullable(UInt64)) AS value_pct, 0 AS ord
    FROM avg_all
    UNION ALL
    SELECT '일평균', '평일', avg_cnt, CAST(NULL AS Nullable(UInt64)), 1
    FROM avg_dayType WHERE day_type='weekday'
    UNION ALL
    SELECT '일평균', '주말', avg_cnt, CAST(NULL AS Nullable(UInt64)), 2
    FROM avg_dayType WHERE day_type='weekend'
    UNION ALL
    SELECT '성별경향', gender, avg_cnt,
           toUInt64(round(avg_cnt / (SELECT cnt FROM overall_cnt) * 100)) AS value_pct,
           10 + if(gender='남성',0,1) AS ord
    FROM avg_gender WHERE gender IN ('남성','여성')
    UNION ALL
    SELECT '연령대경향',
           concat(toString(rk),'위_',age_group)                         AS label,
           avg_cnt,
           toUInt64(round(avg_cnt / (SELECT cnt FROM overall_cnt) * 100)) AS value_pct,
           20 + rk AS ord
    FROM rank_age WHERE rk<=3
    UNION ALL
    SELECT '시간대경향',
           concat('평일_',toString(rk),'_',time_range)                 AS label,
           visit_cnt, pct, 30 + rk
    FROM rank_slot WHERE day_type='weekday' AND rk<=3
    UNION ALL
    SELECT '시간대경향',
           concat('주말_',toString(rk),'_',time_range),
           visit_cnt, pct, 40 + rk
    FROM rank_slot WHERE day_type='weekend' AND rk<=3
)
SELECT section, label, value_cnt, value_pct
FROM final
ORDER BY ord
"""

        answer = ""
        # 요청된 매장들 처리 (더미데이터 포함)
        store_names = [state["store_name"]] if isinstance(state["store_name"], str) else state["store_name"]
        
        for store in store_names:
            # 더미데이터점들인 경우 가짜 데이터 생성
            if store.startswith("더미데이터점"):
                answer += self._generate_dummy_data_for_store(store)
                continue
                
            # 실제 매장인 경우 DB 조회
            database = 'plusinsight'  # 기본 데이터베이스
            try:
                # 특정 데이터베이스로 클라이언트 재생성
                store_client = self._create_clickhouse_client(database=database)
                result = store_client.query(query)

                if len(result.result_rows) > 0:
                    # 섹션별로 데이터 분류
                    sections = {
                        '일평균': [],
                        '성별경향': [],
                        '연령대경향': [],
                        '시간대경향': []
                    }
                    
                    for row in result.result_rows:
                        section, label, value_cnt, value_pct = row
                        sections[section].append((label, value_cnt, value_pct))
                    
                    # 표 형태로 포맷팅 (mcp_diagnose.py와 동일한 형식)
                    store_answer = f"\n=== {store} ===\n"
                    
                    # 1. 일평균 방문객수
                    store_answer += "일평균 방문객수\n"
                    for label, cnt, _ in sections['일평균']:
                        store_answer += f"  {label}: {cnt}명\n"
                    
                    # 2. 성별경향
                    store_answer += "\n성별경향\n"
                    for label, cnt, pct in sections['성별경향']:
                        gender_display = 'M' if label == '남성' else 'F'
                        store_answer += f"  {gender_display}: {pct}%\n"
                    
                    # 3. 연령대별 순위 (상위 3개)
                    store_answer += "\n연령대별 순위\n"
                    for label, cnt, pct in sections['연령대경향']:
                        rank = label.split('위_')[0]
                        age_group = label.split('위_')[1]
                        store_answer += f"  {rank}: {age_group} - {pct}%\n"
                    
                    # 4. 주요 방문시간대
                    store_answer += "\n주요 방문시간대\n"
                    
                    # 시간대 명칭 매핑
                    time_names = {
                        '22-01': '심야',
                        '02-05': '새벽',
                        '06-09': '아침',
                        '10-13': '낮',
                        '14-17': '오후',
                        '18-21': '저녁'
                    }
                    
                    # 평일 시간대 분리
                    weekday_slots = [item for item in sections['시간대경향'] if '평일_' in item[0]]
                    weekend_slots = [item for item in sections['시간대경향'] if '주말_' in item[0]]
                    
                    store_answer += "  평일:\n"
                    for label, cnt, pct in weekday_slots:
                        rank = label.split('_')[1]
                        time_range = label.split('_')[2]
                        time_name = time_names.get(time_range, time_range)
                        store_answer += f"    {rank}: {time_name}({time_range}) - {pct}%\n"
                    
                    store_answer += "  주말:\n"
                    for label, cnt, pct in weekend_slots:
                        rank = label.split('_')[1]
                        time_range = label.split('_')[2]
                        time_name = time_names.get(time_range, time_range)
                        store_answer += f"    {rank}: {time_name}({time_range}) - {pct}%\n"
                    
                    answer += store_answer
                else:
                    answer += f"\n{store} 데이터가 없습니다."
                
            except Exception as e:
                self.logger.error(f"{store} 데이터 조회 오류: {e}")
                answer += f"\n{store} 데이터 조회 오류: {e}"

        state["raw_answer"] = answer
        return state

    def _create_clickhouse_client(self, database="plusinsight"):
        """ClickHouse 클라이언트 생성 (mcp_diagnose.py의 로직을 그대로 사용)"""
        import clickhouse_connect
        from dotenv import load_dotenv
        
        load_dotenv()
        
        # 환경변수 로드
        CLICKHOUSE_HOST = os.getenv("CLICKHOUSE_HOST")
        CLICKHOUSE_PORT = os.getenv("CLICKHOUSE_PORT")
        CLICKHOUSE_USER = os.getenv("CLICKHOUSE_USER")
        CLICKHOUSE_PASSWORD = os.getenv("CLICKHOUSE_PASSWORD")
        SSH_HOST = os.getenv("SSH_HOST")
        SSH_PORT = int(os.getenv("SSH_PORT", "22"))
        SSH_USERNAME = os.getenv("SSH_USERNAME")
        SSH_PASSWORD = os.getenv("SSH_PASSWORD")
        
        # SSH 터널링이 필요한 경우 (sshtunnel 사용)
        try:
            from sshtunnel import SSHTunnelForwarder
            SSH_AVAILABLE = True
        except ImportError:
            SSH_AVAILABLE = False
            
        if SSH_AVAILABLE and SSH_HOST:
            try:
                ssh_tunnel = SSHTunnelForwarder(
                    (SSH_HOST, SSH_PORT),
                    ssh_username=SSH_USERNAME,
                    ssh_password=SSH_PASSWORD,
                    remote_bind_address=(CLICKHOUSE_HOST, int(CLICKHOUSE_PORT)),
                    local_bind_address=("localhost", 0),
                )
                ssh_tunnel.start()
                self.logger.info(f"SSH 터널 생성: localhost:{ssh_tunnel.local_bind_port}")
                
                host = "localhost"
                port = ssh_tunnel.local_bind_port
                
            except Exception as e:
                self.logger.error(f"SSH 터널 생성 실패: {e}, 직접 연결 시도")
                host = CLICKHOUSE_HOST
                port = int(CLICKHOUSE_PORT)
        else:
            # 직접 연결
            host = CLICKHOUSE_HOST
            port = int(CLICKHOUSE_PORT)
        
        try:
            client = clickhouse_connect.get_client(
                host=host,
                port=port,
                username=CLICKHOUSE_USER,
                password=CLICKHOUSE_PASSWORD,
                database=database,
            )
            self.logger.info(f"ClickHouse 연결 성공: {host}:{port}, db={database}")
            return client
        except Exception as e:
            self.logger.error(f"ClickHouse 연결 실패: {e}")
            return None

    def _parse_node(self, state: VisitorDiagnoseState) -> VisitorDiagnoseState:
        """raw_text → metric_dict - 모든 데이터 파싱"""
        import re

        self.logger.info(f"파싱할 raw_answer 길이: {len(state['raw_answer'])}")
        
        metric_dict = {}
        current_store = None
        current_section = None
        
        for line in state["raw_answer"].splitlines():
            line = line.strip()
            if not line:
                continue
                
            # 매장명 파싱
            m = re.match(r"===\s(.+)\s===", line)
            if m:
                current_store = m.group(1)
                metric_dict[current_store] = {
                    'daily_avg': {},      # 일평균 방문객수
                    'gender': {},         # 성별경향  
                    'age_rank': {},       # 연령대별 순위
                    'time_slots': {}      # 주요 방문시간대
                }
                self.logger.info(f"매장 발견: {current_store}")
                continue
            
            # 섹션 헤더 파싱
            if line in ['일평균 방문객수', '성별경향', '연령대별 순위', '주요 방문시간대']:
                current_section = line
                continue
                
            if not current_store or not current_section:
                continue
                
            # 각 섹션별 데이터 파싱
            if current_section == '일평균 방문객수':
                m = re.match(r"(.+):\s+(\d+)명", line)
                if m:
                    label, count = m.groups()
                    metric_dict[current_store]['daily_avg'][label] = int(count)
                    
            elif current_section == '성별경향':
                m = re.match(r"([MF]):\s+(\d+)%", line)
                if m:
                    gender, pct = m.groups()
                    gender_name = '남성' if gender == 'M' else '여성'
                    metric_dict[current_store]['gender'][gender_name] = int(pct)
                    
            elif current_section == '연령대별 순위':
                m = re.match(r"(\d+):\s+(.+)\s+-\s+(\d+)%", line)
                if m:
                    rank, age_group, pct = m.groups()
                    metric_dict[current_store]['age_rank'][f"{rank}위_{age_group}"] = int(pct)
                    
            elif current_section == '주요 방문시간대':
                # 평일/주말 구분
                if line in ['평일:', '주말:']:
                    current_day_type = line.rstrip(':')
                    if current_day_type not in metric_dict[current_store]['time_slots']:
                        metric_dict[current_store]['time_slots'][current_day_type] = {}
                    continue
                    
                # 시간대 데이터 파싱
                m = re.match(r"(\d+):\s+(.+)\((.+)\)\s+-\s+(\d+)%", line)
                if m and 'current_day_type' in locals():
                    rank, time_name, time_range, pct = m.groups()
                    key = f"{rank}위_{time_name}_{time_range}"
                    metric_dict[current_store]['time_slots'][current_day_type][key] = int(pct)

        state["metric_dict"] = metric_dict
        self.logger.info(f"파싱 완료 - 매장 수: {len(metric_dict)}")
        for store, data in metric_dict.items():
            self.logger.info(f"{store} 파싱 결과:")
            self.logger.info(f"  일평균: {data['daily_avg']}")
            self.logger.info(f"  성별: {data['gender']}")
            self.logger.info(f"  연령대: {data['age_rank']}")
            self.logger.info(f"  시간대: {data['time_slots']}")
        return state

    def _map_to_excel_node(self, state: VisitorDiagnoseState) -> VisitorDiagnoseState:
        """
        사용자 수정 템플릿에 맞는 DataFrame 구조로 생성
        매장별 4컬럼: {매장}_템플릿1, {매장}_결과, {매장}_템플릿2, {매장}_참고값
        """
        import pandas as pd
        import os
        import numpy as np
        
        metric_dict = state["metric_dict"]
        excel_path = "report/점포진단표.xlsx"
        
        # report 디렉토리 생성
        if not os.path.exists("report"):
            os.makedirs("report", exist_ok=True)
        
        # 기존 엑셀 파일이 있으면 삭제하고 새로 만들기
        if os.path.exists(excel_path):
            os.remove(excel_path)
            self.logger.info(f"기존 엑셀 파일 삭제: {excel_path}")
        
        # 기본 고정 템플릿 구조 (A~C 컬럼)
        rows = [
            {"주제": "방문객수", "항목": "일평균", "지표": "평균 방문객 수"},
            {"주제": "방문객수", "항목": "일평균", "지표": "평일 방문객 수"},
            {"주제": "방문객수", "항목": "일평균", "지표": "주말 방문객 수"},
            {"주제": "방문객수", "항목": "성별경향", "지표": "남성"},
            {"주제": "방문객수", "항목": "성별경향", "지표": "여성"},
            {"주제": "방문객수", "항목": "연령대 경향", "지표": "연령대 별 순위"},
            {"주제": "방문객수", "항목": "연령대 경향", "지표": "연령대 별 순위"},
            {"주제": "방문객수", "항목": "연령대 경향", "지표": "연령대 별 순위"},
            {"주제": "방문객수", "항목": "시간대 경향", "지표": "평일 주요 방문시간대"},
            {"주제": "방문객수", "항목": "시간대 경향", "지표": "평일 주요 방문시간대"},
            {"주제": "방문객수", "항목": "시간대 경향", "지표": "평일 주요 방문시간대"},
            {"주제": "방문객수", "항목": "시간대 경향", "지표": "주말 주요 방문시간대"},
            {"주제": "방문객수", "항목": "시간대 경향", "지표": "주말 주요 방문시간대"},
            {"주제": "방문객수", "항목": "시간대 경향", "지표": "주말 주요 방문시간대"},
        ]
        
        df = pd.DataFrame(rows)
        self.logger.info(f"기본 템플릿 DataFrame 생성: {len(df)}행")
        
        # 매장별 데이터 추가 (매장당 4컬럼)
        for store, data in metric_dict.items():
            self.logger.info(f"매장 {store} 데이터 추가 중")
            
            # 템플릿1 고정값들
            template1_values = [
                np.nan,  # 평균 방문객 수
                "평일",   # 평일 방문객 수
                "주말",   # 주말 방문객 수
                "M",     # 남성
                "F",     # 여성
                "1",     # 1위 연령대
                "2",     # 2위 연령대
                "3",     # 3위 연령대
                "1",     # 평일 1위 시간대
                "2",     # 평일 2위 시간대
                "3",     # 평일 3위 시간대
                "1",     # 주말 1위 시간대
                "2",     # 주말 2위 시간대
                "3",     # 주말 3위 시간대
            ]
            
            # 결과 데이터 (실제 측정값)
            result_values = []
            
            # 일평균 방문객수
            daily_avg = data.get('daily_avg', {})
            result_values.extend([
                daily_avg.get('전체', 0),
                daily_avg.get('평일', 0),
                daily_avg.get('주말', 0)
            ])
            
            # 성별경향
            gender = data.get('gender', {})
            result_values.extend([
                gender.get('남성', 0),
                gender.get('여성', 0)
            ])
            
            # 연령대별 순위 (상위 3개) - 실제 연령대 값 추출
            age_rank = data.get('age_rank', {})
            age_values = []
            for i in range(1, 4):
                rank_key = None
                for key in age_rank.keys():
                    if key.startswith(f"{i}위_"):
                        rank_key = key
                        break
                if rank_key:
                    # "1위_50대" → "50" 추출
                    age_group = rank_key.split('위_')[1].replace('대', '')
                    age_values.append(int(age_group) if age_group.isdigit() else 0)
                else:
                    age_values.append(0)
            result_values.extend(age_values)
            
            # 시간대 데이터 변환: "1위_저녁_18-21" → "저녁(18-21)"
            time_slots = data.get('time_slots', {})
            
            # 평일 시간대
            weekday_slots = time_slots.get('평일', {})
            for i in range(1, 4):
                rank_key = None
                for key in weekday_slots.keys():
                    if key.startswith(f"{i}위_"):
                        rank_key = key
                        break
                if rank_key:
                    # "1위_저녁_18-21" → "저녁(18-21)"
                    parts = rank_key.split('_')
                    if len(parts) >= 3:
                        time_name = parts[1]  # 저녁
                        time_range = parts[2]  # 18-21
                        formatted_time = f"{time_name}({time_range})"
                        result_values.append(formatted_time)
                    else:
                        result_values.append(np.nan)
                else:
                    result_values.append(np.nan)
                    
            # 주말 시간대
            weekend_slots = time_slots.get('주말', {})
            for i in range(1, 4):
                rank_key = None
                for key in weekend_slots.keys():
                    if key.startswith(f"{i}위_"):
                        rank_key = key
                        break
                if rank_key:
                    # "1위_저녁_18-21" → "저녁(18-21)"
                    parts = rank_key.split('_')
                    if len(parts) >= 3:
                        time_name = parts[1]  # 저녁
                        time_range = parts[2]  # 18-21
                        formatted_time = f"{time_name}({time_range})"
                        result_values.append(formatted_time)
                    else:
                        result_values.append(np.nan)
                else:
                    result_values.append(np.nan)
            
            # 템플릿2 고정값들 (단위)
            template2_values = [
                "명",    # 평균 방문객 수
                "명",    # 평일 방문객 수
                "명",    # 주말 방문객 수
                "%",     # 남성
                "%",     # 여성
                "대",    # 1위 연령대
                "대",    # 2위 연령대
                "대",    # 3위 연령대
                np.nan,  # 평일 1위 시간대
                np.nan,  # 평일 2위 시간대
                np.nan,  # 평일 3위 시간대
                np.nan,  # 주말 1위 시간대
                np.nan,  # 주말 2위 시간대
                np.nan,  # 주말 3위 시간대
            ]
            
            # 참고값 (현재는 연령대와 시간대에만 비율값)
            reference_values = []
            
            # 일평균, 성별 - NaN
            reference_values.extend([np.nan] * 5)
            
            # 연령대 비율값
            for i in range(1, 4):
                rank_key = None
                for key in age_rank.keys():
                    if key.startswith(f"{i}위_"):
                        rank_key = key
                        break
                pct_value = age_rank.get(rank_key, 0) if rank_key else 0
                reference_values.append(f"{pct_value}%" if pct_value > 0 else np.nan)
            
            # 시간대 비율값
            for day_type in ['평일', '주말']:
                slots = time_slots.get(day_type, {})
                for i in range(1, 4):
                    rank_key = None
                    for key in slots.keys():
                        if key.startswith(f"{i}위_"):
                            rank_key = key
                            break
                    pct_value = slots.get(rank_key, 0) if rank_key else 0
                    reference_values.append(f"{pct_value}%" if pct_value > 0 else np.nan)
            
            # 매장별 4컬럼 추가
            df[f"{store}_템플릿1"] = template1_values
            df[f"{store}_결과"] = result_values
            df[f"{store}_템플릿2"] = template2_values
            df[f"{store}_참고값"] = reference_values
            
            self.logger.info(f"{store} 4컬럼 추가 완료: 결과 {len(result_values)}개 값")
        
        # 엑셀 파일로 저장 (헤더 없이 저장)
        df.to_excel(excel_path, sheet_name="방문객분석", index=False, header=False)
        
        # 수동으로 올바른 헤더 추가
        import openpyxl
        from openpyxl.styles import Alignment
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['방문객분석']
        
        # 모든 데이터를 두 행씩 아래로 이동 (헤더 2행 확보)
        ws.insert_rows(1, 2)
        
        # 첫 번째 행 A1:C1 은 빈칸으로 유지(지점명 표시 제거)
        ws.merge_cells('A1:C1')  # 3칸 병합 (값 없음)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 매장별 4개 컬럼을 매장명으로 병합 (첫 번째 행)
        col_start = 4  # D 컬럼부터 시작
        for store in metric_dict.keys():
            start_col = openpyxl.utils.get_column_letter(col_start)
            end_col = openpyxl.utils.get_column_letter(col_start + 3)
            merge_range = f"{start_col}1:{end_col}1"
            
            ws.merge_cells(merge_range)
            ws[f"{start_col}1"].value = store
            ws[f"{start_col}1"].alignment = Alignment(horizontal='center', vertical='center')
            
            col_start += 4  # 다음 매장으로 4컬럼씩 이동
        
        # 두 번째 행: 주제, 항목, 지표, 결과, 참고값 등
        ws.cell(2, 1).value = "주제"
        ws.cell(2, 2).value = "항목" 
        ws.cell(2, 3).value = "지표"

        # 매장별로 결과(3셀 병합), 참고값 헤더
        col_start = 4
        for store in metric_dict.keys():
            # 결과 3셀 병합
            start_col = openpyxl.utils.get_column_letter(col_start)
            end_col = openpyxl.utils.get_column_letter(col_start + 2)
            ws.merge_cells(f"{start_col}2:{end_col}2")
            ws[f"{start_col}2"].value = "결과"
            ws[f"{start_col}2"].alignment = Alignment(horizontal='center', vertical='center')

            # 참고값 헤더
            ref_col = openpyxl.utils.get_column_letter(col_start + 3)
            ws[f"{ref_col}2"].value = "참고값"
            ws[f"{ref_col}2"].alignment = Alignment(horizontal='center', vertical='center')

            col_start += 4
        
        wb.save(excel_path)
        
        self.logger.info(f"템플릿 DataFrame을 엑셀로 저장: {excel_path}")
        self.logger.info(f"DataFrame 형태: {df.shape}")
        
        # DataFrame 내용 로그 출력 (처음 5행만)
        self.logger.info("DataFrame 샘플 (처음 5행):")
        self.logger.info(f"\n{df.head().to_string()}")
        
        # placements는 빈 리스트로 설정
        state["placements"] = []
        state["dataframe"] = df
        
        return state

    def _update_excel_node(self, state: VisitorDiagnoseState) -> VisitorDiagnoseState:
        """
        pandas DataFrame으로 이미 엑셀 저장이 완료되었으므로 상태만 업데이트
        추가로 셀 병합과 정렬 적용
        """
        # placements가 비어있으면 pandas로 이미 저장 완료
        if not state["placements"]:
            self.logger.info("pandas DataFrame으로 엑셀 저장 완료")
            
            # 셀 병합과 정렬 적용
            excel_path = "report/점포진단표.xlsx"
            self._apply_cell_merge_and_alignment(excel_path)
            
            state["final_result"] = "엑셀 업데이트 완료 (pandas + 셀병합)"
            return state
        
        # 혹시 placements가 있다면 기존 로직 실행
        import asyncio
        
        async def update_excel():
            from mcp_use import MCPClient  # tools import 제거
            import os

            config = {
                "mcpServers": {
                    "excel": {"command": "npx", "args": ["--yes", "@negokaz/excel-mcp-server"]}
                }
            }
            client = MCPClient.from_dict(config)
            await client.create_all_sessions()
            session = client.get_session("excel")

            excel_file = os.path.abspath("report/점포진단표.xlsx")
            for p in state["placements"]:
                await session.connector.call_tool(
                    "excel_write_to_sheet",
                    {
                        "fileAbsolutePath": excel_file,
                        "sheetName": p["sheet"],
                        "newSheet": False,
                        "range": p["cell"],
                        "values": [[p["value"]]],
                    },
                )
            await client.close_all_sessions()
        
        # 비동기 함수를 동기적으로 실행
        asyncio.run(update_excel())
        state["final_result"] = "엑셀 업데이트 완료 (mcp-server)"
        return state

    def _apply_cell_merge_and_alignment(self, excel_path: str):
        """
        셀 병합, 정렬, 테두리, 크기를 단계별로 적용합니다.
        """
        self.logger.info("셀 병합 및 정렬 적용 시작")
        self._apply_cell_merging(excel_path)
        self._apply_alignment(excel_path)
        self._apply_borders(excel_path)
        self._apply_cell_sizing(excel_path)
        self.logger.info(f"셀 병합 및 정렬 적용 완료: {excel_path}")

    def _apply_cell_merging(self, excel_path: str):
        """1단계: 셀 병합만 처리"""
        from openpyxl import load_workbook
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        self.logger.info("1단계: 셀 병합 적용")
        
        # 주제 컬럼 (A) - 모든 방문객수 병합
        ws.merge_cells('A3:A16')
        ws['A3'].value = "방문객수"
        
        # 항목 컬럼 (B) - 같은 내용끼리 병합
        merge_ranges = [
            ('B3:B5', '일평균'),
            ('B6:B7', '성별경향'),
            ('B8:B10', '연령대 경향'),
            ('B11:B16', '시간대 경향')
        ]
        
        for range_str, value in merge_ranges:
            ws.merge_cells(range_str)
            first_cell = range_str.split(':')[0]
            ws[first_cell].value = value
        
        # 지표 컬럼 (C) - 같은 내용끼리 병합
        ws.merge_cells('C8:C10')
        ws['C8'].value = "연령대 별 순위"
        ws.merge_cells('C11:C13')
        ws['C11'].value = "평일 주요 방문시간대"
        ws.merge_cells('C14:C16')
        ws['C14'].value = "주말 주요 방문시간대"
        
        # 참고값 컬럼 빈칸 세로 병합 (일평균/성별만) - 하드코딩
        # 4개 매장 기준: G, K, O, S 컬럼
        for col_letter in ['G', 'K', 'O', 'S']:
            ws.merge_cells(f'{col_letter}3:{col_letter}5')  # 일평균
            ws.merge_cells(f'{col_letter}6:{col_letter}7')  # 성별경향
        
        wb.save(excel_path)
        self.logger.info("1단계 완료: 셀 병합")

    def _apply_alignment(self, excel_path: str):
        """2단계: 정렬만 처리"""
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        self.logger.info("2단계: 정렬 적용")
        
        # 정렬 스타일 정의
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # 첫 번째 행 (매장명) - 모두 중앙 정렬
        for col in range(1, ws.max_column + 1):
            ws.cell(1, col).alignment = center_alignment
        
        # 두 번째 행 (헤더) - 중앙 정렬
        for col in range(1, ws.max_column + 1):
            ws.cell(2, col).alignment = center_alignment
        
        # 데이터 행들 정렬
        for row in range(3, ws.max_row + 1):
            # A열 (주제) - 중앙 정렬
            ws.cell(row, 1).alignment = center_alignment
            # B열 (항목) - 중앙 정렬  
            ws.cell(row, 2).alignment = center_alignment
            # C열 (지표) - 좌측 정렬
            ws.cell(row, 3).alignment = left_alignment
            
            # 매장 데이터 컬럼들 (D~S, 4개 매장 × 4컬럼)
            for col in range(4, ws.max_column + 1):
                col_type = (col - 4) % 4  # 0:템플릿1, 1:결과, 2:템플릿2, 3:참고값
                if col_type == 0:  # 템플릿1 - 중앙 정렬
                    ws.cell(row, col).alignment = center_alignment
                elif col_type == 1:  # 결과 - 우측 정렬 (숫자)
                    ws.cell(row, col).alignment = right_alignment
                elif col_type == 2:  # 템플릿2 - 좌측 정렬 (단위)
                    ws.cell(row, col).alignment = left_alignment
                elif col_type == 3:  # 참고값 - 우측 정렬
                    ws.cell(row, col).alignment = right_alignment
        
        wb.save(excel_path)
        self.logger.info("2단계 완료: 정렬")

    def _apply_borders(self, excel_path: str):
        """3단계: 테두리만 처리 (excel-mcp-server 사용)"""
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side
        
        self.logger.info("3단계: 테두리 적용")
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 테두리 스타일 정의
        thin_border = Side(style='thin')
        medium_border = Side(style='medium')
        thick_border = Side(style='thick')
        
        # 샘플 파일 기반 하드코딩된 테두리 패턴
        # 첫 번째 행 (매장명 헤더)
        for col in range(4, ws.max_column + 1, 4):  # D, H, L, P (매장 시작 컬럼)
            # 매장명 첫 컬럼
            ws.cell(1, col).border = Border(left=medium_border, right=medium_border, top=medium_border)
            # 매장명 중간 컬럼들
            for i in range(1, 3):
                ws.cell(1, col + i).border = Border(top=medium_border)
            # 매장명 마지막 컬럼
            ws.cell(1, col + 3).border = Border(right=medium_border, top=medium_border)
        
        # 두 번째 행 (주제/항목/지표/결과/참고값 헤더)
        ws.cell(2, 1).border = Border(left=medium_border, right=thin_border, top=medium_border, bottom=medium_border)  # 주제
        ws.cell(2, 2).border = Border(left=thin_border, right=thin_border, top=medium_border, bottom=medium_border)   # 항목
        ws.cell(2, 3).border = Border(left=thin_border, top=medium_border, bottom=medium_border)                     # 지표
        
        # 매장별 결과/참고값 헤더
        for col in range(4, ws.max_column + 1, 4):
            # 결과 (3컬럼 병합의 첫 컬럼)
            ws.cell(2, col).border = Border(left=thick_border, right=thin_border, top=medium_border, bottom=medium_border)
            ws.cell(2, col + 1).border = Border(top=medium_border, bottom=medium_border)
            ws.cell(2, col + 2).border = Border(right=thin_border, top=medium_border, bottom=medium_border)
            # 참고값
            ws.cell(2, col + 3).border = Border(left=thin_border, top=medium_border, bottom=medium_border)
        
        # 데이터 행들 (3-16행)
        for row in range(3, 17):
            # A열 (주제) - 세로 병합된 셀
            ws.cell(row, 1).border = Border(left=medium_border, right=thin_border)
            if row == 16:  # 마지막 행
                ws.cell(row, 1).border = Border(left=medium_border, right=thin_border, bottom=medium_border)
            
            # B열 (항목)
            if row in [3, 6, 8, 11]:  # 항목 시작 행들
                ws.cell(row, 2).border = Border(right=thin_border, top=medium_border if row in [6, 11] else thin_border)
            elif row in [5, 7, 10, 16]:  # 항목 끝 행들
                ws.cell(row, 2).border = Border(right=thin_border, bottom=medium_border if row in [5, 7, 16] else thin_border)
            else:
                ws.cell(row, 2).border = Border(right=thin_border)
            
            # C열 (지표)
            if row <= 7:  # 일평균, 성별경향
                if row in [3, 4, 5]:  # 일평균
                    ws.cell(row, 3).border = Border(left=thin_border, bottom=thin_border if row == 5 else None)
                elif row in [6, 7]:  # 성별경향
                    ws.cell(row, 3).border = Border(left=thin_border, top=medium_border if row == 6 else None, bottom=medium_border if row == 7 else thin_border)
            elif row in [8, 9, 10]:  # 연령대
                ws.cell(row, 3).border = Border(left=thin_border, right=thick_border, top=thin_border if row == 8 else None, bottom=thin_border if row == 10 else None)
            elif row in [11, 12, 13]:  # 평일 시간대
                ws.cell(row, 3).border = Border(left=thin_border, right=thick_border, top=medium_border if row == 11 else None, bottom=thin_border if row == 13 else None)
            elif row in [14, 15, 16]:  # 주말 시간대
                ws.cell(row, 3).border = Border(left=thin_border, right=thick_border, top=thin_border if row == 14 else None, bottom=medium_border if row == 16 else None)
            
            # 매장 데이터 컬럼들 (D부터)
            for col in range(4, ws.max_column + 1):
                col_type = (col - 4) % 4  # 0:템플릿1, 1:결과, 2:템플릿2, 3:참고값
                
                # 기본 테두리
                left_border = thick_border if col_type == 0 else None
                right_border = None
                top_border = None
                bottom_border = None
                
                # 행별 특별 테두리
                if row == 3:  # 첫 데이터 행
                    top_border = None
                    bottom_border = thin_border
                elif row in [4, 5]:  # 일평균 나머지
                    bottom_border = thin_border if row == 4 else None
                elif row in [6, 7]:  # 성별경향
                    top_border = medium_border if row == 6 else None
                    bottom_border = medium_border if row == 7 else thin_border
                elif row in [8, 9, 10]:  # 연령대
                    bottom_border = thin_border
                elif row in [11, 12, 13]:  # 평일 시간대
                    top_border = medium_border if row == 11 else None
                    bottom_border = thin_border
                elif row in [14, 15, 16]:  # 주말 시간대
                    top_border = thin_border if row == 14 else None
                    bottom_border = medium_border if row == 16 else thin_border
                
                # 참고값 컬럼 (마지막 컬럼)
                if col_type == 3:
                    left_border = thin_border
                
                ws.cell(row, col).border = Border(
                    left=left_border,
                    right=right_border,
                    top=top_border,
                    bottom=bottom_border
                )
        
        wb.save(excel_path)
        self.logger.info("3단계 완료: 테두리 (하드코딩)")

    def _apply_cell_sizing(self, excel_path: str):
        """4단계: 셀 크기만 처리"""
        from openpyxl import load_workbook
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        self.logger.info("4단계: 셀 크기 적용")
        
        # 행 높이 조정
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 25
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 12  # 주제
        ws.column_dimensions['B'].width = 18  # 항목
        ws.column_dimensions['C'].width = 15  # 지표
        
        # 매장 컬럼들 너비 조정
        for col_idx in range(4, ws.max_column + 1):
            col_letter = chr(ord('A') + col_idx - 1)
            ws.column_dimensions[col_letter].width = 12
        
        wb.save(excel_path)
        self.logger.info("4단계 완료: 셀 크기")

    def _generate_dummy_data_for_store(self, store_name: str) -> str:
        """더미데이터점들을 위한 가짜 데이터 생성 (각 매장마다 다른 특성)"""
        self.logger.info(f"{store_name} 더미 데이터 생성")
        
        # 매장별로 다른 더미 데이터 패턴
        dummy_patterns = {
            "더미데이터점": {
                "daily": {"전체": 380, "평일": 395, "주말": 350},
                "gender": {"M": 52, "F": 48},
                "age": [("40대", 28), ("50대", 22), ("30대", 20)],
                "time_weekday": [("오후(14-17)", 26), ("저녁(18-21)", 24), ("낮(10-13)", 19)],
                "time_weekend": [("저녁(18-21)", 27), ("오후(14-17)", 23), ("낮(10-13)", 22)]
            },
            "더미데이터점1": {
                "daily": {"전체": 320, "평일": 310, "주말": 340},
                "gender": {"M": 48, "F": 52},
                "age": [("30대", 32), ("20대", 26), ("40대", 24)],
                "time_weekday": [("저녁(18-21)", 28), ("오후(14-17)", 25), ("아침(06-09)", 20)],
                "time_weekend": [("오후(14-17)", 30), ("저녁(18-21)", 26), ("낮(10-13)", 24)]
            },
            "더미데이터점2": {
                "daily": {"전체": 450, "평일": 470, "주말": 420},
                "gender": {"M": 45, "F": 55},
                "age": [("50대", 35), ("60대이상", 28), ("40대", 25)],
                "time_weekday": [("낮(10-13)", 30), ("오후(14-17)", 27), ("저녁(18-21)", 22)],
                "time_weekend": [("낮(10-13)", 32), ("오후(14-17)", 28), ("아침(06-09)", 21)]
            },
            "더미데이터점3": {
                "daily": {"전체": 280, "평일": 260, "주말": 320},
                "gender": {"M": 58, "F": 42},
                "age": [("20대", 38), ("10대", 30), ("30대", 22)],
                "time_weekday": [("아침(06-09)", 25), ("낮(10-13)", 23), ("오후(14-17)", 21)],
                "time_weekend": [("저녁(18-21)", 35), ("심야(22-01)", 28), ("오후(14-17)", 25)]
            }
        }
        
        # 해당 매장의 패턴 가져오기 (없으면 기본 더미데이터점 패턴 사용)
        pattern = dummy_patterns.get(store_name, dummy_patterns["더미데이터점"])
        
        # 더미 데이터 생성
        dummy_data = f"""
=== {store_name} ===
일평균 방문객수
  전체: {pattern['daily']['전체']}명
  평일: {pattern['daily']['평일']}명
  주말: {pattern['daily']['주말']}명

성별경향
  M: {pattern['gender']['M']}%
  F: {pattern['gender']['F']}%

연령대별 순위
  1: {pattern['age'][0][0]} - {pattern['age'][0][1]}%
  2: {pattern['age'][1][0]} - {pattern['age'][1][1]}%
  3: {pattern['age'][2][0]} - {pattern['age'][2][1]}%

주요 방문시간대
  평일:
    1: {pattern['time_weekday'][0][0]} - {pattern['time_weekday'][0][1]}%
    2: {pattern['time_weekday'][1][0]} - {pattern['time_weekday'][1][1]}%
    3: {pattern['time_weekday'][2][0]} - {pattern['time_weekday'][2][1]}%
  주말:
    1: {pattern['time_weekend'][0][0]} - {pattern['time_weekend'][0][1]}%
    2: {pattern['time_weekend'][1][0]} - {pattern['time_weekend'][1][1]}%
    3: {pattern['time_weekend'][2][0]} - {pattern['time_weekend'][2][1]}%
"""
        return dummy_data


# FastMCP 인스턴스 (툴 서버 등록용)
mcp = FastMCP("visitor_diagnose_excel")


@mcp.tool()  # FastMCP 서버 전용
def visitor_diagnose_excel(
    *,
    store_name: Union[str, List[str]],
    start_date: str,
    end_date: str,
    user_prompt: str = "매장 방문객 진단 분석 엑셀화"
) -> str:
    """[EXCEL_REPORT] Generate an **Excel report** for *visitor diagnostics*.

    Trigger words (case-insensitive):
        - "엑셀", "excel", "xlsx", "엑셀화", "sheet", "보고서"
        - Combinations like "방문객 진단 엑셀", "visitor diagnose excel" etc.

    Use this when the user explicitly asks to *export/produce an Excel file* of
    visitor-related metrics such as daily average visitors, gender ratio,
    age ranking, or time-slot trends.

    Parameters
    ----------
    store_name : str | list[str]
        Store name(s) to diagnose. Accepts a single string or a list.
    start_date : str
        Start date (YYYY-MM-DD).
    end_date : str
        End date (YYYY-MM-DD).
    user_prompt : str, optional
        Custom prompt for LLM. Defaults to "매장 방문객 진단 분석 엑셀화".

    Returns
    -------
    str
        Result message containing "엑셀 업데이트 완료" when successful.
    """

    workflow = VisitorDiagnoseWorkflow()
    return workflow.run(
        user_prompt=user_prompt,
        store_name=store_name,
        start_date=start_date,
        end_date=end_date,
    )


if __name__ == "__main__":
    # FastMCP 서버 실행
    print("FastMCP 서버 시작 - visitor_diagnose", file=sys.stderr)
    try:
        mcp.run()
    except Exception as e:
        print(f"서버 오류 발생: {e}", file=sys.stderr)
